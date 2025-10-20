#!/usr/bin/env python3
"""
BYTE Panel - Template Creator for Google Sheets
Genera archivo Excel con estructura completa del panel de control
Para importar a Google Sheets
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️  Instalando dependencia: openpyxl")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Colores BYTE
BYTE_BLUE = "0066CC"
BYTE_DARK = "003366"
BYTE_WHITE = "FFFFFF"

# Colores de estado
COLOR_PADRE = "E3F2FD"
COLOR_INDIVIDUAL = "FFF3E0"
COLOR_ACTIVO = "C8E6C9"
COLOR_INACTIVO = "FFCDD2"
COLOR_MODIFICADO = "FFF9C4"

def create_header_style():
    """Estilo para headers"""
    return {
        'font': Font(name='Arial', size=11, bold=True, color=BYTE_WHITE),
        'fill': PatternFill(start_color=BYTE_DARK, end_color=BYTE_DARK, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_title_style():
    """Estilo para título principal"""
    return {
        'font': Font(name='Arial', size=18, bold=True, color=BYTE_WHITE),
        'fill': PatternFill(start_color=BYTE_BLUE, end_color=BYTE_BLUE, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }

def apply_style(cell, style_dict):
    """Aplicar estilo a celda"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def create_productos_sheet(wb):
    """Crear hoja de Productos con estructura completa"""
    ws = wb.active
    ws.title = "Productos"

    # Headers
    headers = [
        "ID Producto",
        "ID Variante",
        "Handle",
        "SKU",
        "Tipo",
        "Título",
        "Marca",
        "Modelo",
        "Descripción",
        "Precio",
        "Precio Comparación",
        "Stock",
        "# Variantes",
        "Estado",
        "Categoría",
        "Proveedor",
        "URL Imagen",
        "Última Sync",
        "Modificado",
        "Sincronizar"
    ]

    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        apply_style(cell, create_header_style())

    # Ajustar ancho de columnas
    column_widths = {
        'A': 12,  # ID Producto
        'B': 12,  # ID Variante
        'C': 25,  # Handle
        'D': 20,  # SKU
        'E': 12,  # Tipo
        'F': 35,  # Título
        'G': 15,  # Marca
        'H': 20,  # Modelo
        'I': 40,  # Descripción
        'J': 12,  # Precio
        'K': 18,  # Precio Comparación
        'L': 10,  # Stock
        'M': 12,  # # Variantes
        'N': 12,  # Estado
        'O': 15,  # Categoría
        'P': 15,  # Proveedor
        'Q': 30,  # URL Imagen
        'R': 18,  # Última Sync
        'S': 12,  # Modificado
        'T': 12,  # Sincronizar
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freezar primera fila
    ws.freeze_panes = 'A2'

    # Agregar filas de ejemplo
    ejemplos = [
        [
            "123456789",
            "987654321",
            "michelin-ps4-205-40-17-p",
            "MICH-PS4-205-40-17",
            "PADRE",
            "Michelin PILOT SPORT 4 205/40R17",
            "Michelin",
            "PILOT SPORT 4",
            "Neumático de alto rendimiento para uso deportivo",
            197471,
            "",
            20,
            10,
            "ACTIVO",
            "Neumáticos",
            "Michelin",
            "https://cdn.shopify.com/...",
            "",
            "FALSE",
            "FALSE"
        ],
        [
            "234567890",
            "876543210",
            "bfg-ko2-265-70-17-i",
            "BFG-KO2-265-70-17",
            "INDIVIDUAL",
            "BFGoodrich All-Terrain T/A KO2 265/70R17",
            "BFGoodrich",
            "All-Terrain T/A KO2",
            "Neumático todo terreno con tecnología CoreGard",
            185000,
            "",
            15,
            1,
            "ACTIVO",
            "Neumáticos",
            "BFGoodrich",
            "https://cdn.shopify.com/...",
            "",
            "FALSE",
            "FALSE"
        ]
    ]

    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col_num, value in enumerate(ejemplo, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='center')

            # Color según tipo
            if col_num == 5:  # Tipo
                if value == "PADRE":
                    cell.fill = PatternFill(start_color=COLOR_PADRE, end_color=COLOR_PADRE, fill_type='solid')
                elif value == "INDIVIDUAL":
                    cell.fill = PatternFill(start_color=COLOR_INDIVIDUAL, end_color=COLOR_INDIVIDUAL, fill_type='solid')

    # Agregar notas/comentarios
    ws.cell(row=1, column=6).comment = "✏️ EDITABLE - Puedes modificar este campo"
    ws.cell(row=1, column=9).comment = "✏️ EDITABLE - Puedes modificar este campo"
    ws.cell(row=1, column=10).comment = "✏️ EDITABLE - Formato: número sin símbolos"
    ws.cell(row=1, column=11).comment = "✏️ EDITABLE - Precio tachado (opcional)"
    ws.cell(row=1, column=12).comment = "✏️ EDITABLE - Solo números enteros"
    ws.cell(row=1, column=14).comment = "✏️ EDITABLE - ACTIVO o INACTIVO"

    return ws

def create_dashboard_sheet(wb):
    """Crear hoja de Dashboard"""
    ws = wb.create_sheet("Dashboard")

    # Título principal
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "🚀 BYTE - Panel de Control Shopify"
    apply_style(title_cell, create_title_style())
    ws.row_dimensions[1].height = 35

    # Subtítulo
    ws.merge_cells('A2:F2')
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Desarrollado por BYTE para Rodavial"
    subtitle_cell.font = Font(name='Arial', size=11, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center')

    # Sección: Estadísticas Generales
    row = 4
    ws.merge_cells(f'A{row}:B{row}')
    section_cell = ws.cell(row=row, column=1)
    section_cell.value = "📊 ESTADÍSTICAS GENERALES"
    section_cell.font = Font(name='Arial', size=12, bold=True, color=BYTE_WHITE)
    section_cell.fill = PatternFill(start_color=BYTE_DARK, end_color=BYTE_DARK, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')

    # Estadísticas
    stats = [
        ("Total de Productos:", "=COUNTA(Productos!A:A)-1"),
        ("Productos Padre:", '=COUNTIF(Productos!E:E,"PADRE")'),
        ("Productos Individuales:", '=COUNTIF(Productos!E:E,"INDIVIDUAL")'),
        ("", ""),
        ("✅ Activos:", '=COUNTIF(Productos!N:N,"ACTIVO")'),
        ("❌ Inactivos:", '=COUNTIF(Productos!N:N,"INACTIVO")'),
        ("", ""),
        ("⚠️ Modificados (pendientes):", '=COUNTIF(Productos!S:S,TRUE)'),
    ]

    row += 1
    for label, formula in stats:
        row += 1
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row, column=2).value = formula
        ws.cell(row=row, column=2).font = Font(name='Arial', size=10)

    # Sección: Inventario
    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    section_cell = ws.cell(row=row, column=1)
    section_cell.value = "📦 INVENTARIO"
    section_cell.font = Font(name='Arial', size=12, bold=True, color=BYTE_WHITE)
    section_cell.fill = PatternFill(start_color=BYTE_DARK, end_color=BYTE_DARK, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')

    inventory_stats = [
        ("Stock Total:", "=SUM(Productos!L:L)"),
        ("Valor Total:", '=SUMPRODUCT(Productos!L:L,Productos!J:J)'),
    ]

    row += 1
    for label, formula in inventory_stats:
        row += 1
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row, column=2).value = formula
        ws.cell(row=row, column=2).font = Font(name='Arial', size=10)
        if "Valor" in label:
            ws.cell(row=row, column=2).number_format = '$#,##0.00'

    # Sección: TOP Marcas
    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    section_cell = ws.cell(row=row, column=1)
    section_cell.value = "🏆 TOP MARCAS"
    section_cell.font = Font(name='Arial', size=12, bold=True, color=BYTE_WHITE)
    section_cell.fill = PatternFill(start_color=BYTE_DARK, end_color=BYTE_DARK, fill_type='solid')
    section_cell.alignment = Alignment(horizontal='center')

    row += 1
    ws.cell(row=row, column=1).value = "Marca"
    ws.cell(row=row, column=2).value = "Cantidad"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)

    # Ejemplos de marcas (en producción se calcularía dinámicamente)
    marcas_ejemplo = [
        ("Michelin", '=COUNTIF(Productos!G:G,"Michelin")'),
        ("BFGoodrich", '=COUNTIF(Productos!G:G,"BFGoodrich")'),
    ]

    for marca, formula in marcas_ejemplo:
        row += 1
        ws.cell(row=row, column=1).value = marca
        ws.cell(row=row, column=2).value = formula

    # Footer
    row += 3
    ws.merge_cells(f'A{row}:B{row}')
    footer_cell = ws.cell(row=row, column=1)
    footer_cell.value = "Última actualización: Se actualiza automáticamente con cada cambio"
    footer_cell.font = Font(name='Arial', size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal='center')

    row += 2
    ws.merge_cells(f'A{row}:B{row}')
    footer_cell = ws.cell(row=row, column=1)
    footer_cell.value = "Powered by BYTE 🚀"
    footer_cell.font = Font(name='Arial', size=10, bold=True, color=BYTE_BLUE)
    footer_cell.alignment = Alignment(horizontal='center')

    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    return ws

def create_historial_sheet(wb):
    """Crear hoja de Historial"""
    ws = wb.create_sheet("Historial")

    # Título
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "📜 Historial de Cambios"
    apply_style(title_cell, create_title_style())
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        "Fecha/Hora",
        "Usuario",
        "Handle",
        "Campo",
        "Valor Anterior",
        "Valor Nuevo",
        "Fila"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        apply_style(cell, create_header_style())

    # Anchos
    column_widths = {
        'A': 20,  # Fecha/Hora
        'B': 25,  # Usuario
        'C': 25,  # Handle
        'D': 20,  # Campo
        'E': 25,  # Valor Anterior
        'F': 25,  # Valor Nuevo
        'G': 10,  # Fila
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freezar headers
    ws.freeze_panes = 'A3'

    # Ejemplo de registro
    ejemplo = [
        "2025-01-17 14:30:22",
        "usuario@rodavial.com",
        "michelin-ps4-205-40-17-p",
        "Precio",
        "197471",
        "199500",
        "2"
    ]

    for col_num, value in enumerate(ejemplo, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = value
        cell.alignment = Alignment(vertical='center')

    # Nota informativa
    ws.cell(row=5, column=1).value = "ℹ️ Este historial se genera automáticamente al editar productos en la hoja 'Productos'"
    ws.merge_cells('A5:G5')
    ws.cell(row=5, column=1).font = Font(italic=True, color="666666")

    return ws

def create_ayuda_sheet(wb):
    """Crear hoja de Ayuda"""
    ws = wb.create_sheet("Ayuda")

    # Título
    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "❓ Guía de Uso - BYTE Panel de Control"
    apply_style(title_cell, create_title_style())
    ws.row_dimensions[1].height = 30

    # Contenido de ayuda
    help_content = [
        ("", ""),
        ("🎯 CAMPOS EDITABLES", ""),
        ("", ""),
        ("✅ Puedes editar:", "❌ NO editar (solo lectura):"),
        ("• Título", "• ID Producto"),
        ("• Descripción", "• ID Variante"),
        ("• Precio", "• Handle"),
        ("• Precio Comparación", "• SKU"),
        ("• Stock", "• Tipo"),
        ("• Estado (ACTIVO/INACTIVO)", "• Marca"),
        ("", "• Modelo"),
        ("", "• # Variantes"),
        ("", "• URL Imagen"),
        ("", "• Última Sync"),
        ("", "• Modificado (automático)"),
        ("", "• Sincronizar (automático)"),
        ("", ""),
        ("📝 CÓMO EDITAR PRODUCTOS", ""),
        ("", ""),
        ("1. Ve a la hoja 'Productos'", ""),
        ("2. Busca el producto (Ctrl+F)", ""),
        ("3. Edita el campo que necesites", ""),
        ("4. El sistema marcará automáticamente como 'Modificado'", ""),
        ("5. En máximo 15 minutos se sincronizará con Shopify", ""),
        ("", ""),
        ("⚠️ VALIDACIONES IMPORTANTES", ""),
        ("", ""),
        ("• Precio: Solo números, sin símbolos ($, puntos, comas)", "Correcto: 197471 | Incorrecto: $197.471"),
        ("• Stock: Solo números enteros positivos", "Correcto: 20 | Incorrecto: 20.5 o -5"),
        ("• Estado: Exactamente 'ACTIVO' o 'INACTIVO'", "Correcto: ACTIVO | Incorrecto: activo o Active"),
        ("", ""),
        ("📊 DASHBOARD", ""),
        ("", ""),
        ("• Ve a la hoja 'Dashboard' para ver:", ""),
        ("  - Total de productos", ""),
        ("  - Productos activos/inactivos", ""),
        ("  - Stock total", ""),
        ("  - Valor del inventario", ""),
        ("  - Top marcas", ""),
        ("", ""),
        ("🔄 SINCRONIZACIÓN", ""),
        ("", ""),
        ("• Automática cada 15 minutos", ""),
        ("• Bidireccional: Shopify ↔ Google Sheets", ""),
        ("• Solo se actualizan productos modificados", ""),
        ("", ""),
        ("📞 SOPORTE BYTE", ""),
        ("", ""),
        ("📧 Email: contacto@byte.com.ar", ""),
        ("📱 WhatsApp: +54 9 11 XXXX-XXXX", ""),
        ("🌐 Web: www.byte.com.ar", ""),
        ("", ""),
        ("Horario: Lunes a Viernes 9:00-18:00 | Sábados 9:00-13:00", ""),
    ]

    row = 2
    for line in help_content:
        row += 1
        ws.cell(row=row, column=1).value = line[0]
        if len(line) > 1:
            ws.cell(row=row, column=3).value = line[1]

        # Estilo para títulos de sección
        if line[0].startswith("🎯") or line[0].startswith("📝") or line[0].startswith("⚠️") or line[0].startswith("📊") or line[0].startswith("🔄") or line[0].startswith("📞"):
            ws.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True, color=BYTE_BLUE)
            ws.merge_cells(f'A{row}:D{row}')

    # Anchos
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 5

    return ws

def main():
    """Crear archivo Excel con todas las hojas"""
    print("🚀 BYTE Panel - Generador de Template")
    print("=" * 50)
    print()

    # Crear workbook
    print("📝 Creando archivo Excel...")
    wb = Workbook()

    # Crear hojas
    print("📄 Creando hoja 'Productos'...")
    create_productos_sheet(wb)

    print("📊 Creando hoja 'Dashboard'...")
    create_dashboard_sheet(wb)

    print("📜 Creando hoja 'Historial'...")
    create_historial_sheet(wb)

    print("❓ Creando hoja 'Ayuda'...")
    create_ayuda_sheet(wb)

    # Guardar
    filename = "BYTE_Panel_Template.xlsx"
    print(f"💾 Guardando archivo: {filename}...")
    wb.save(filename)

    print()
    print("✅ ¡Template creado exitosamente!")
    print()
    print("📋 PRÓXIMOS PASOS:")
    print()
    print("1. Subir este archivo a Google Drive")
    print("2. Click derecho → Abrir con → Google Sheets")
    print("3. Google Sheets lo convertirá automáticamente")
    print("4. Instalar Google Apps Script (ver documentación)")
    print("5. Configurar n8n workflow")
    print()
    print("📖 Ver documentación completa en:")
    print("   Documentacion/BYTE_CONTROL_PANEL.md")
    print()
    print("🚀 Powered by BYTE")
    print("=" * 50)

if __name__ == "__main__":
    main()
